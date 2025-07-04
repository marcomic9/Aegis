import openpyxl
from datetime import datetime
import os
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Configure logging for debugging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

def create_excel_file(scheme_number, data, output_file=None):
    try:
        # Validate inputs
        if not isinstance(scheme_number, str):
            raise ValueError("scheme_number must be a string")
        if not isinstance(data, list):
            raise ValueError("data must be a list of dictionaries")

        # Generate dynamic filename if not provided
        if output_file is None:
            # Replace spaces and special characters with underscores for filename safety
            safe_scheme_number = scheme_number.replace(" ", "_").replace(":", "_")
            current_date = datetime.now().strftime("%Y-%m-%d_%I-%M-%p-SAST")
            output_file = f"{safe_scheme_number}_{current_date}.xlsx"

        logging.info(f"Creating Excel file: {output_file}")
        logging.info(f"Current working directory: {os.getcwd()}")

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None  # Address Pylance warning
        ws.title = "Aegis Report"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        title_font = Font(bold=True, size=14)
        cell_font = Font(size=11)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        center_align = Alignment(horizontal="center", wrap_text=True)
        padding_align = Alignment(horizontal="center", wrap_text=True)

        # Get the current system date and time (05:52 PM SAST, 2025-06-07)
        current_date = datetime.now().strftime("%Y-%m-%d %I:%M %p SAST")

        # Write and format the report date in cell A1, merged across A1:E1
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Report Date: {current_date}"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        ws["A1"].border = thin_border
        ws.row_dimensions[1].height = 25  # Increased height for better spacing

        # Write and format the Sectional Scheme Number in cell A2, merged across A2:E2
        ws.merge_cells("A2:E2")
        ws["A2"] = scheme_number
        ws["A2"].font = title_font
        ws["A2"].alignment = center_align
        ws["A2"].border = thin_border
        ws.row_dimensions[2].height = 25  # Increased height for better spacing

        # Write and format the column headers in row 3
        headers = ["ID", "Name", "Phone 1", "Phone 2", "Phone 3"]  # Shortened headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        ws.row_dimensions[3].height = 22  # Increased height for headers

        # Write the data starting from row 4
        for row_idx, entry in enumerate(data, start=4):
            if not isinstance(entry, dict):
                logging.warning(f"Skipping invalid entry at index {row_idx-4}, expected a dictionary")
                continue
            row_data = [
                entry.get("ID", ""),
                entry.get("Name", ""),
                entry.get("Phone1", ""),
                entry.get("Phone2", ""),
                entry.get("Phone3", "")
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = padding_align
            ws.row_dimensions[row_idx].height = 20  # Increased height for data rows

        # Auto-adjust column widths for columns A to E with minimum and maximum constraints
        for col in range(1, 6):  # Columns A to E
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max(max_length + 4, 10)  # Minimum width 10, add padding
            adjusted_width = min(adjusted_width, 20)  # Maximum width 20 to avoid excessive space
            ws.column_dimensions[column].width = adjusted_width

        # Ensure the output directory exists
        os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)

        # Save the Excel file
        wb.save(output_file)
        logging.info(f"Excel file saved as {os.path.abspath(output_file)}")

    except PermissionError as e:
        logging.error(f"Cannot save file '{output_file}'. Check if the file is open or you lack permissions. Details: {e}")
        raise
    except Exception as e:
        logging.error(f"Failed to create Excel file. Details: {e}")
        raise

# Test script with sample data
if __name__ == "__main__":
    # Define test data
    test_scheme_number = "Test Scheme Number"
    test_data = [
        {"ID": "001", "Name": "Alice", "Phone1": "123456789", "Phone2": "987654321", "Phone3": "555555555"},
        {"ID": "002", "Name": "Bob", "Phone1": "111222333"},
        {"ID": "003", "Name": "Charlie", "Phone1": "444555666", "Phone2": "777888999"}
    ]

    # Run the function with test data
    try:
        create_excel_file(test_scheme_number, test_data)
    except Exception as e:
        logging.error(f"Test failed: {e}")